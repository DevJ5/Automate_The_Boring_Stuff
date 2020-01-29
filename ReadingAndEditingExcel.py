import openpyxl
import os

os.chdir('C:\\Users\\veens\\Downloads')

# Open the workbook
workbook = openpyxl.load_workbook('example.xlsx')

# Open a sheet
sheet = workbook['Sheet1']
# Open multiple sheets
sheets = workbook.sheetnames

# Get a cell
cell = sheet['A1']

# Get cell alternative (rows and columns start at 1 not 0)
cell2 = sheet.cell(row=1, column=1)

print(cell.value, cell2.value)

# Loop through sheet
for i in range (1, 8):
    print(i, sheet.cell(i, 2).value)


# Editing
workbook2 = openpyxl.Workbook()

sheet2 = workbook2['Sheet']
sheet2['A1'].value = "Hello"
sheet2['A2'].value = "World"
# This value only exists in memory, it has to be saved to the file
workbook2.save('mypythonexcel.xlsx')


# Create new sheet
workbook2.create_sheet(index=0, title='myfirstsheet')

# Change title of sheet
sheet2.title = "My new sheet name"

workbook2.save('mypythonexcel.xlsx')