import openpyxl
import os

os.chdir('C:\\Users\\veens\\Downloads')

# Open the workbook
workbook = openpyxl.load_workbook('example.xlsx')

# Open a sheet
sheet = workbook['Sheet1']

# Get a cell
cell = sheet['A1']

# Get cell alternative (rows and columns start at 1 not 0)
cell2 = sheet.cell(row=1, column=1)

print(cell.value, cell2.value)

# Loop through sheet
for i in range (1, 8):
    print(i, sheet.cell(i, 2).value)

sheets = workbook.sheetnames
